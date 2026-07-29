"""Readable Solace exports and reviewed bill-import helpers."""
from __future__ import annotations

import csv
import io
from calendar import monthrange
from datetime import date, datetime, time
from decimal import Decimal, InvalidOperation

from django.utils import timezone
from django.utils.dateparse import parse_date, parse_datetime

from apps.solace import selectors
from apps.solace.serializers import (
    AccountBalanceSnapshotSerializer,
    BillOccurrenceSerializer,
    CycleCloseoutSerializer,
    FinanceCategorySerializer,
    PaydayChecklistItemSerializer,
    SolaceSettingsSerializer,
    SubscriptionSerializer,
)


def _yes(value) -> str:
    return "yes" if value else "no"


def bill_rows(user) -> list[dict]:
    return [
        {
            "name": row.name,
            "amount": f"{row.amount:.2f}",
            "category": row.category,
            "provider": row.provider,
            "due_at": row.due_at.isoformat() if row.due_at else "",
            "recurrence_rule": row.recurrence_rule,
            "active": _yes(row.is_active),
            "include_in_set_aside": _yes(row.include_in_set_aside),
            "notes": row.notes,
        }
        for row in selectors.list_bills(user)
    ]


def purchase_rows(user) -> list[dict]:
    return [
        {
            "name": row.name,
            "target_amount": f"{row.target_amount:.2f}",
            "amount_saved": f"{row.saved_amount:.2f}",
            "target_date": row.target_date.isoformat() if row.target_date else "",
            "category": row.category,
            "priority": row.priority,
            "status": row.status,
            "notes": row.notes,
        }
        for row in selectors.list_purchases(user)
    ]


def income_rows(user) -> list[dict]:
    return [
        {
            "name": row.title,
            "amount": f"{row.expected_amount:.2f}",
            "next_pay_date": row.pay_at.isoformat() if row.pay_at else "",
            "recurrence_rule": row.recurrence_rule,
            "active": _yes(row.is_active),
            "notes": row.notes,
        }
        for row in selectors.list_paydays(user)
    ]


def bucket_rows(user) -> list[dict]:
    return [
        {
            "name": row.name,
            "category": row.category,
            "target_amount": f"{row.target_amount:.2f}",
            "current_amount": f"{row.current_amount:.2f}",
            "allocation_method": row.allocation_method,
            "allocation_value": f"{row.allocation_value:.2f}",
            "rounding_increment": f"{row.rounding_increment:.2f}",
            "cap_to_remaining": _yes(row.cap_to_remaining),
            "active": _yes(row.is_active),
            "sort_order": row.position,
            "notes": row.notes,
        }
        for row in selectors.list_buckets(user)
    ]


def export_sheets(user) -> list[tuple[str, list[dict]]]:
    bills = selectors.list_bills(user)
    occurrences = []
    for bill in bills:
        occurrences.extend(list(bill.occurrences.order_by("due_at")))
    settings_obj = selectors.get_settings()
    from apps.solace.models import CycleCloseout

    closeouts = list(CycleCloseout.objects.order_by("-cycle_start"))
    return [
        ("Bills", bill_rows(user)),
        ("Bill Occurrences", list(BillOccurrenceSerializer(occurrences, many=True).data)),
        ("Planned Purchases", purchase_rows(user)),
        ("Income Sources", income_rows(user)),
        ("Buckets", bucket_rows(user)),
        ("Subscriptions", list(SubscriptionSerializer(selectors.list_subscriptions(user), many=True).data)),
        ("Account Balances", list(AccountBalanceSnapshotSerializer(selectors.list_balance_snapshots(user), many=True).data)),
        ("Categories", list(FinanceCategorySerializer(selectors.list_categories(user), many=True).data)),
        ("Checklist", list(PaydayChecklistItemSerializer(selectors.list_checklist_items(user), many=True).data)),
        ("Closeouts", list(CycleCloseoutSerializer(closeouts, many=True).data)),
        ("Settings", [dict(SolaceSettingsSerializer(settings_obj).data)] if settings_obj else []),
    ]


def csv_bytes(rows: list[dict]) -> bytes:
    output = io.StringIO()
    if rows:
        writer = csv.DictWriter(output, fieldnames=list(rows[0].keys()))
        writer.writeheader()
        writer.writerows(
            {
                key: _safe_spreadsheet_cell(value)
                for key, value in row.items()
            }
            for row in rows
        )
    return output.getvalue().encode("utf-8-sig")


def xlsx_bytes(user) -> bytes:
    from openpyxl import Workbook

    workbook = Workbook()
    workbook.remove(workbook.active)
    for title, rows in export_sheets(user):
        sheet = workbook.create_sheet(title[:31])
        if not rows:
            continue
        headers = list(rows[0].keys())
        sheet.append(headers)
        for row in rows:
            sheet.append(
                [
                    _safe_spreadsheet_cell(
                        value.isoformat() if isinstance(value, (date, datetime)) else value
                    )
                    for value in (row.get(header, "") for header in headers)
                ]
            )
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def _safe_spreadsheet_cell(value):
    if isinstance(value, str) and value.startswith(("=", "+", "-", "@")):
        return f"'{value}"
    return value


def read_uploaded_rows(upload) -> list[dict]:
    raw = upload.read()
    filename = (upload.name or "").lower()
    if filename.endswith(".csv"):
        return list(csv.DictReader(io.StringIO(raw.decode("utf-8-sig"))))
    if filename.endswith(".xlsx"):
        from openpyxl import load_workbook

        workbook = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        sheet = workbook.active
        iterator = sheet.iter_rows(values_only=True)
        headers = [
            _normal_key(value)
            for value in (next(iterator, ()) or ())
        ]
        rows = []
        for values in iterator:
            row = {
                header: value
                for header, value in zip(headers, values, strict=False)
                if header
            }
            if any(value not in (None, "") for value in row.values()):
                rows.append(row)
        return rows
    raise ValueError("Upload a CSV or XLSX file.")


def _normal_key(value) -> str:
    return str(value or "").strip().lower().replace(" ", "_")


def _pick(row: dict, *names: str, default=""):
    values = {_normal_key(key): value for key, value in row.items()}
    for name in names:
        value = values.get(_normal_key(name))
        if value not in (None, ""):
            return value
    return default


def _bool(value, default=True) -> bool:
    if value in (None, ""):
        return default
    return str(value).strip().lower() not in {"no", "false", "0", "off"}


_FREQUENCIES = {
    "weekly": "FREQ=WEEKLY",
    "fortnightly": "FREQ=WEEKLY;INTERVAL=2",
    "monthly": "FREQ=MONTHLY",
    "quarterly": "FREQ=MONTHLY;INTERVAL=3",
    "six-monthly": "FREQ=MONTHLY;INTERVAL=6",
    "six monthly": "FREQ=MONTHLY;INTERVAL=6",
    "yearly": "FREQ=YEARLY",
    "one-off": "",
    "one off": "",
}


def _due_at(row: dict, frequency: str) -> datetime | None:
    explicit = str(_pick(row, "due_at", "due_date", default="")).strip()
    if explicit:
        parsed_datetime = parse_datetime(explicit)
        if parsed_datetime:
            return (
                timezone.make_aware(parsed_datetime, timezone.get_current_timezone())
                if timezone.is_naive(parsed_datetime)
                else parsed_datetime
            )
        parsed_date = parse_date(explicit)
        if parsed_date:
            return timezone.make_aware(
                datetime.combine(parsed_date, time(hour=9)),
                timezone.get_current_timezone(),
            )
        raise ValueError("Due date must use ISO format.")

    start_value = _pick(row, "start_date", "start", default="")
    start = parse_date(str(start_value)) if start_value else timezone.localdate()
    if not start:
        raise ValueError("Start date must use YYYY-MM-DD.")
    due_day = int(Decimal(str(_pick(row, "due_day", "day", "due", default=start.day))))
    if not 1 <= due_day <= 31:
        raise ValueError("Due day must be 1–31.")
    month = start.month
    due_month_value = _pick(row, "due_month", "month", default="")
    if due_month_value not in (None, ""):
        month = int(Decimal(str(due_month_value)))
        if not 1 <= month <= 12:
            raise ValueError("Due month must be 1–12.")
    day = min(due_day, monthrange(start.year, month)[1])
    return timezone.make_aware(
        datetime.combine(date(start.year, month, day), time(hour=9)),
        timezone.get_current_timezone(),
    )


def parse_bill_import_rows(rows: list[dict]) -> tuple[list[dict], int]:
    parsed = []
    error_count = 0
    for source_row, row in enumerate(rows, start=2):
        preview = {"source_row": source_row, "errors": []}
        try:
            name = str(_pick(row, "name", "bill", "bill_name")).strip()
            if not name:
                preview["errors"].append("Missing name")
            try:
                amount = Decimal(str(_pick(row, "amount", "cost", default="0"))).quantize(
                    Decimal("0.01")
                )
                if amount <= 0:
                    preview["errors"].append("Amount must be greater than zero")
            except InvalidOperation:
                amount = Decimal("0.00")
                preview["errors"].append("Amount is not a number")
            recurrence = str(_pick(row, "recurrence_rule", default="")).strip()
            frequency = str(_pick(row, "frequency", default="monthly")).strip().lower()
            if not recurrence:
                if frequency not in _FREQUENCIES:
                    preview["errors"].append(f"Unsupported frequency: {frequency}")
                recurrence = _FREQUENCIES.get(frequency, "")
            due_at = _due_at(row, frequency)
            preview.update(
                {
                    "name": name,
                    "amount": f"{amount:.2f}",
                    "category": str(_pick(row, "category", default="other")).strip() or "other",
                    "provider": str(_pick(row, "provider", "account_name", "account", default="")).strip(),
                    "due_at": due_at.isoformat() if due_at else None,
                    "recurrence_rule": recurrence,
                    "is_active": _bool(_pick(row, "is_active", "active", default="yes")),
                    "include_in_set_aside": _bool(
                        _pick(row, "include_in_set_aside", "include", default="yes")
                    ),
                    "notes": str(_pick(row, "notes", default="")).strip(),
                }
            )
        except (TypeError, ValueError) as exc:
            preview["errors"].append(str(exc))
        if preview["errors"]:
            error_count += 1
        parsed.append(preview)
    return parsed, error_count
